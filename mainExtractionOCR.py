import cv2
import os
import pytesseract
from matplotlib import pyplot as pt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import Border, Side
from drawingNum import GetString

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

wb = Workbook()

for image in range(1,21):
    img_path = os.path.join("images", f'{image:02}.png')
    
    # Check if the image file exists before reading
    if (os.path.exists(img_path) == False):
        print("Image ", f'{image:02}.png', " not found.")
        continue
        
    init_img = cv2.imread(img_path, 0)
    [init_row, init_col] = init_img.shape
    
    
    # --- Cropping + border image --- #
    img = init_img[12:init_row-15, 12:init_col-12]
    [nrow, ncol] = img.shape
    
    
    # --- Isolating vertical & horizontal lines --- #
    ret, bin_img= cv2.threshold(img, 127, 255, cv2.THRESH_BINARY_INV)
    
    horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, ncol//150))
    eroded_verti = cv2.erode(bin_img, horiz_kernel, iterations = 5)
    vertical_lines = cv2.dilate(eroded_verti, horiz_kernel, iterations = 5)
    
    verti_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (nrow//150, 1))
    eroded_hori = cv2.erode(bin_img, verti_kernel, iterations=5)
    horizontal_lines = cv2.dilate(eroded_hori, verti_kernel, iterations = 5)
    
    combined_lines = cv2.bitwise_or(vertical_lines, horizontal_lines)
    
    
    # --- Drawing remove --- #
    rect_kernel3 = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    drawingMask = cv2.erode(combined_lines, rect_kernel3, iterations = 2)
    drawingMask = cv2.dilate(drawingMask, rect_kernel3, iterations = 50)
    table_lines = drawingMask + np.bitwise_not(combined_lines)
    
    
    # --- Removing arrow lines --- #
    table_lines_dil = cv2.dilate(np.bitwise_not(table_lines), rect_kernel3, iterations = 5)
    
    contours, hierarchy = cv2.findContours(table_lines_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    sorted_contours = sorted(contours, key=cv2.contourArea, reverse = False)
    
    
    # --- Filling remaining drawing contours w/ white --- #
    table_bgr = cv2.cvtColor(table_lines, cv2.COLOR_GRAY2BGR)
    
    for i in range(0, len(sorted_contours)):
        cntr = sorted_contours[i]
        x, y, w, h = cv2.boundingRect(cntr)
        if (w < 30 or h < 30):
            cv2.drawContours(table_bgr, sorted_contours, i, (255, 255, 255), thickness=-1)
    
    table_only = cv2.cvtColor(table_bgr, cv2.COLOR_BGR2GRAY)
    _, table_only = cv2.threshold(table_only, 150, 255, cv2.THRESH_BINARY)
            
    
    # --- Isolating table cells --- #
    table_only_copy = cv2.copyMakeBorder(table_only, 5, 5, 5, 5, cv2.BORDER_CONSTANT, 0)
    table_lines_dil2 = cv2.dilate(np.bitwise_not(table_only_copy), rect_kernel3, iterations = 1)
    cell_cntr, hierarchy = cv2.findContours(table_lines_dil2, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    
    # --- Creating mask for tables + obtaining coordinates for useful cells --- #
    table_bgr2 = cv2.cvtColor(table_only, cv2.COLOR_GRAY2BGR)
    
    keywords = ["DRAWING NUMBER", "DRAWING NO", "DRAWN BY", "DRAWN", "CHECKED BY", "CHECKED", "TITLE", "DRAWING TITLE", "APPPROVED BY", "APPROVED", "CONTRACTOR", "COMPANY", "UNIT", "STATUS", "PAGE", "PROJECT NO", "PROJECT NUM", "LANG", "CAD NO", "FONT", "FONT STYLE", "AMENDMENTS"]
    useful_cells = []
    
    for c in cell_cntr:
        coordinates = cv2.boundingRect(c)
        x, y, w, h = coordinates
        rect_area = w * h
        if (rect_area < ((nrow//4) * (ncol//4)) and h < 400):
            cell = img[y:y+h, x:x+w]
            string = (pytesseract.image_to_string(cell, config ='--psm 6')).strip()
            string_list = string.splitlines()
            for k in keywords:
                if k in string:
                    cell_info = [k, coordinates, string_list]
                    useful_cells.append(cell_info)
                    
            
            # --- Masking tables --- #
            cv2.rectangle(table_bgr2, (x, y), (x+w, y+h), (0, 0, 0), -1)
    
    table_mask = cv2.cvtColor(table_bgr2, cv2.COLOR_BGR2GRAY)
    table_mask = cv2.dilate(np.bitwise_not(table_mask), rect_kernel3, iterations=5)
    
    drawing = np.bitwise_not(bin_img) + table_mask
    drawing[drawing >= 5] = 255
    drawing[drawing < 5] = 0
    
    tables = np.bitwise_not(bin_img) + np.bitwise_not(table_mask)
    tables[tables >= 5] = 255
    tables[tables < 5] = 0
    
    
    # --- Checking for unattended full-vertical tables --- #
    _, bin_drawing = cv2.threshold(drawing, 150, 255, cv2.THRESH_BINARY_INV)
    bin_drawing = cv2.erode(bin_drawing, horiz_kernel, iterations = 5)
    bin_drawing = cv2.dilate(bin_drawing, horiz_kernel, iterations = 5)
    
    
    vertical_lines_dil = cv2.dilate(bin_drawing, rect_kernel3, iterations = 2)
    vert_contours, hierarchy = cv2.findContours(vertical_lines_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    vert_tf = False
    for c in vert_contours:
        x, y, w, h = cv2.boundingRect(c)
        if (h >= nrow - 50):
            vert_tf = True
            break
        
    # --- Extracting with largest contour --- #
    if (vert_tf == True and len(useful_cells) == 0):
        drawing_mask2 = np.zeros((nrow, ncol), dtype=np.uint8)
        
        contours, _ = cv2.findContours(np.bitwise_not(bin_img), cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
        sorted_contours = sorted(contours, key=cv2.contourArea, reverse = True)
        x, y, w, h = cv2.boundingRect(sorted_contours[0])
        drawing_mask2[y:y+h, x:x+w] = 255
        
        tables = np.bitwise_not(bin_img) + drawing_mask2
        tables[tables >= 5] = 255
        tables[tables < 5] = 0
        
        drawing = np.bitwise_not(bin_img) + np.bitwise_not(drawing_mask2)
        drawing[drawing >= 5] = 255
        drawing[drawing < 5] = 0
        
        # Extracting drawing number
        drawingNum = GetString(init_img, "DRAWING NUMBER", "DRAWING NO")
        drawnBy = GetString(init_img, "DRAWN BY",'DRAWN')
        
        if (len(drawingNum) > 0):
            useful_cells.append(["DRAWING NUMBER", None, ["", drawingNum]])
    
    # --- Reanalyzing containing only titles --- #
    for index, info in enumerate(useful_cells):
        if ((len("".join(info[2])) < len(info[0]) + 3)):
            x, y, w, h = info[1]
        
            y_range = 0
            if (info[0] != "AMENDMENTS"):
                y_range = y+h+80
            else:
                y_range = y+h+300
    
            if (y_range > nrow):
                y_range = nrow
            cell = img[y:y_range, x:x+w]
            
            _, cell_thresh = cv2.threshold(cell, 150, 255, cv2.THRESH_BINARY_INV)
            
            # Remove horizontal lines
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40,1))
            remove_horizontal = cv2.morphologyEx(cell_thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            cnts = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = cnts[0] if len(cnts) == 2 else cnts[1]
            for c in cnts:
                cv2.drawContours(cell, [c], -1, (255,255,255), 5)
            
            # Remove vertical lines
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,30))
            remove_vertical = cv2.morphologyEx(cell_thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
            cnts = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = cnts[0] if len(cnts) == 2 else cnts[1]
            for c in cnts:
                cv2.drawContours(cell, [c], -1, (255,255,255), 5)
            
            string = (pytesseract.image_to_string(cell, config ='--psm 6')).strip()
            string_list = string.splitlines()
            info[2] = string_list
            useful_cells[index] = info
    
    
    # --- Sorting and sending data to Excel file --- #
    
    # Removing duplicate cell data
    table_data = []
    
    for c in useful_cells:
        if c in table_data:
            continue
        else:
            table_data.append(c)

    def takeKeyword(list):
        return list[0]
    
    def inList(word, list):
        indexList = []
        for index, elem in enumerate(list):
            for w in word:
                if (w == elem[0]):
                    indexList.append(index)
        return indexList
    
    # Sorting titles in alphabetical order
    table_data.sort(key=takeKeyword)
    
    # Popping Amendments table data out of list
    amend_index = inList(["AMENDMENTS"], table_data)
    amendments = None
    
    if (len(amend_index) == 1):
        amendments = table_data.pop(amend_index[0])
    elif (len(amend_index) > 1): 
        removed = 0
        for i in amend_index:
            if (len(table_data[i - removed][2][1]) > 3):
                amendments = table_data.pop(i - removed)
            else:
                del table_data[i - removed]
            removed += 1
    
    # Removing redundant data (e.g. "DRAWN" & "DRAWN BY")
    def removeExtra(redundant, keep, table_data):
        keywords = []
        for cell in table_data:
            keywords.append(cell[0])
        
        if (keep in keywords):
            if (redundant in keywords):
                del table_data[keywords.index(redundant)]
    
    removeExtra("CHECKED", "CHECKED BY", table_data)
    removeExtra("DRAWN", "DRAWN BY", table_data)
    removeExtra("TITLE", "DRAWING TITLE", table_data)
    removeExtra("APPROVED", "APPROVED BY", table_data)
    
    
    # --- Writing table data into .xlsx file --- #
    
    ws = wb.create_sheet(f'{image:02}.png',image)
    ws.append(["Field Title", "Content"])
    
    thick_border = Border(left=Side(style='thick'),
                         right=Side(style='thick'),
                         top=Side(style='thick'),
                         bottom=Side(style='thick'))
    ws.cell(row=1,column=1).border = thick_border
    ws.cell(row=1,column=2).border = thick_border
    
    for info in table_data:
        ws.append([info[0], info[2][1]])
    
    if (amendments != None):
        title = amendments[2][1].split()
        title_len = len(title)
    
        title_info1 = amendments[2][2].split() 
        title_info2 = amendments[2][3].split()
        
        # --- Creating a table for amendments info --- #
        
        ws['E1'] = "Amendments"
        
        letters = ['E','F','G','H','I']
        if title_len > 3:
            tab = Table(displayName="Amendments"+str(image), ref="E2:I4") 
        elif title_len == 3:
            tab = Table(displayName="Amendments"+str(image), ref="E2:G4") 
        
        for i in range(0,3):
            for x in range(0,title_len):
                ws[letters[x]+'2'] = title[x]
        
        for i in range(0,3):
            for x in range(0,title_len):
                ws[letters[x]+'3'] = title_info1[x]
        
        for i in range(0,3):
            for x in range(0,title_len):
                ws[letters[x]+'4'] = title_info2[x]
        
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    wb.save(filename = 'drawingInfo.xlsx')
    
    # Check if the output folder exists, if not, create it
    writeFolder = "extracted"
    if not os.path.exists(writeFolder):
        os.makedirs(writeFolder)

    # Save the image inside the subfolder
    output_image_path = os.path.join(writeFolder, f'drawing{image:02}.png')
    cv2.imwrite(output_image_path, drawing)
    
    print("Image ", f'{image:02}.png', " processed.")
    
# # --- Display extracted images --- #        
# pt.figure()

# pt.subplot(1, 3, 1)
# pt.title("Original image")
# pt.imshow(img, cmap="gray")

# pt.subplot(1, 3, 2)
# pt.title("Extracted drawing")
# pt.imshow(drawing, cmap="gray")

# pt.subplot(1, 3, 3)
# pt.title("Tables")
# pt.imshow(tables, cmap="gray")

    


